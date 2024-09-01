import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font
from openpyxl.worksheet.pagebreak import Break

# Function to check the length of the given text and shorten it if it exceeds 25 characters
def formats_text(text):
    if len(text) <= 17:
        return text
    else:
        arr_text = text.split(" ")
        print(arr_text)
        formatted_first_word = f'{arr_text[0][0]}.'
        arr_text.pop(0)
        formatted_text = f"{formatted_first_word} {' '.join(arr_text)}"
        return formatted_text

# Load the Excel file and select the active sheet
dataFile = 'ParentInfoList.xlsx'
wb = openpyxl.load_workbook(dataFile)
ws_data = wb.active

# Create a new sheet and configure the page settings
ws_id_cards = wb.create_sheet(title="ID Cards")
ws_id_cards.page_setup.orientation = ws_id_cards.ORIENTATION_PORTRAIT
ws_id_cards.page_setup.paperSize = ws_id_cards.PAPERSIZE_A4
ws_id_cards.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

# Set the font settings
arial_font = Font(name='Arial', size=10)

# Set column widths
ws_id_cards.column_dimensions['A'].width = 19
ws_id_cards.column_dimensions['B'].width = 14
#ws_id_cards.column_dimensions['C'].width = 14
ws_id_cards.column_dimensions['D'].width = 2
ws_id_cards.column_dimensions['E'].width = 19
ws_id_cards.column_dimensions['F'].width = 14
#ws_id_cards.column_dimensions['G'].width = 14

# Initialize starting row and column
start_row = 1
start_col = 0

# Iterate through each row in the data sheet to create ID cards
for index, row in enumerate(ws_data.iter_rows(min_row=2, values_only=True)):
    parent_name, student_name, level_of_student, class_of_student = row

    # Adjust column offset (for switching between cards)
    col_offset = (index % 2) * 4
    current_col = start_col + col_offset

    # Add the top image
    try:
        image_path_top = 'img/BKIDCardFrontTop.png'
        img_top = Image(image_path_top)
        img_top.width = 310
        img_top.height = 80
        ws_id_cards.add_image(img_top, f'{chr(65 + current_col)}{start_row}')
    except FileNotFoundError:
        print(f"{image_path_top} could not be found.")

    # Merge cells and fill in the information
    ws_id_cards.merge_cells(f'{chr(65 + current_col)}{start_row + 1}:{chr(65 + current_col + 2)}{start_row + 4}')
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 5}'] = 'VELİ AD SOYAD'
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 5}'] = f': {formats_text(parent_name)}'
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 6}'] = 'ÖĞRENCİ AD SOYAD'
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 6}'] = f': {formats_text(student_name)}'
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 7}'] = 'KADEME'
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 7}'] = f': {level_of_student}'
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 8}'] = 'SINIF'
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 8}'] = f': {class_of_student}'

    # Apply the font settings
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 5}'].font = arial_font
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 5}'].font = arial_font
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 6}'].font = arial_font
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 6}'].font = arial_font
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 7}'].font = arial_font
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 7}'].font = arial_font
    ws_id_cards[f'{chr(65 + current_col)}{start_row + 8}'].font = arial_font
    ws_id_cards[f'{chr(65 + current_col + 1)}{start_row + 8}'].font = arial_font

    # Add the bottom image
    try:
        image_path_bottom = 'img/BKIDCardFrontBot.png'
        img_bottom = Image(image_path_bottom)
        img_bottom.width = 310
        img_bottom.height = 10
        ws_id_cards.add_image(img_bottom, f'{chr(65 + current_col)}{start_row + 10}')
    except FileNotFoundError:
        print(f"{image_path_bottom} could not be found.")

    # Move to the next row after every two cards
    if index % 2 == 1:
        start_row += 12

    # Add a page break after every 10 cards
    if (index + 1) % 10 == 0 and index != 0:
        ws_id_cards.row_breaks.append(Break(id=start_row - 2))

# Finally, save the new Excel file
wb.save('ParentIDCardList.xlsx')

print('Created ID Cards')
